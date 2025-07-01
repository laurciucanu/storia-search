import os

import requests
from bs4 import BeautifulSoup
import pandas as pd
import time

from openpyxl import load_workbook

BASE_URL = "https://www.storia.ro"
START_URL = "https://www.storia.ro/ro/rezultate/vanzare/apartament/iasi/iasi?limit=72&ownerTypeSingleSelect=ALL&areaMin=65&buildYearMin=1980&roomsNumber=%5BTHREE%2CFOUR%5D&priceMax=160000&by=PRICE&direction=ASC&page={}"

headers = {
    "User-Agent": "Mozilla/5.0"
}

apartamente = []

for page in range(1, 10):
    print(f"Procesare pagina {page}...")
    url = START_URL.format(page)
    r = requests.get(url, headers=headers)
    soup = BeautifulSoup(r.text, 'html.parser')

    articole = soup.find_all("article")
    for art in articole:
        link_tag = art.find("a", href=True)
        if not link_tag or not link_tag["href"].startswith("/ro/oferta/"):
            continue

        link = BASE_URL + link_tag["href"]

        loc_tag = art.find(attrs={"data-sentry-component": "Address"})
        locatie = loc_tag.get_text(strip=True) if loc_tag else "N/A"

        titlu = link_tag["href"].split("/")[-1].replace("-", " ").title()

        pret_tag = art.find(attrs={"data-sentry-component": "Price"})
        pret = pret_tag.get_text(strip=True) if pret_tag else "N/A"

        apartamente.append({
            "Titlu": titlu,
            "Preț": pret,
            "Locație": locatie,
            "Link": link
        })

        apartamente.append({
            "Titlu": titlu,
            "Pret": pret,
            "Locație": locatie,
            "Link": link
        })

    time.sleep(1)

df = pd.DataFrame(apartamente)
df.drop_duplicates(subset="Link", inplace=True)
df.sort_values(by="Pret", inplace=True)
df.drop(columns=["Pret"], inplace=True)

excel_path = "apartamente.xlsx"

def is_file_locked(filepath):
    if not os.path.exists(filepath):
        return False
    try:
        with open(filepath, "r+"):
            return False
    except IOError:
        return True


if is_file_locked(excel_path):
    print(f"⚠️ Fișierul '{excel_path}' este deschis. Închide-l din Excel și rulează din nou.")
    exit(1)

df.to_excel(excel_path, index=False)

wb = load_workbook(excel_path)
ws = wb.active

for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=ws.max_column)
    url = cell.value
    cell.value = "Deschide anunțul"
    cell.hyperlink = url
    cell.style = "Hyperlink"

for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = max_length + 2
    ws.column_dimensions[column].width = adjusted_width

wb.save(excel_path)
print("Salvat în apartamente.xlsx cu coloane auto-ajustate.")
